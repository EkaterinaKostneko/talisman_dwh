'''
Created on 15 Apr 2019

@author: pymancer
'''
import os
import re
import shutil
import pyexcel

from uuid import uuid4
from hashlib import md5
from shutil import copyfile
from collections import defaultdict

from xlrd import XLRDError
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell.cell import MergedCell  # @UnresolvedImport
from df.common.schemas.model import DataType, PIPwareWebEntity, DCDimensionEntity  # @UnresolvedImport
from df.common.helpers.logger import Logger
from df.common.helpers.general import get_prop_by_path  # @UnresolvedImport
from df.common.adapters import get_entity_reader  # @UnresolvedImport
from df.common.uploaders import get_uploader  # @UnresolvedImport
from exceptions import DFExcelExtractorException  # @UnresolvedImport

log = Logger()


class Filepath:
    """ Объект, представляющий путь к файлу загрузки,
        имеющиеся теги данных и их позиции.
        Прототип объекта:
        ('/app/share/20190101-20190131/R_Test_contoso.xlsx', [['20190101', '20190131'], ['contoso']])
    """
    def __init__(self, path, tags=None):
        self.tags = list(tags or tuple())
        self.path = path
        self.f_path = path # original path to disguise possible rename

    def __str__(self):
        return f'Filepath (path -> {self.path}; tags -> {self.tags})'

    def __repr__(self):
        return self.__str__()

    @property
    def path(self):
        return self._path

    @path.setter
    def path(self, path):
        self._path = path
        self.dirname, self.filename = os.path.split(self.path)
        self.basename, self.extension = os.path.splitext(self.filename)

    def append_path(self, path_token):
        self.path = os.path.join(self.path, path_token)
        self.f_path = self.path

    def append_tags(self, tags):
        self.tags.append(list(tags))

    def get_tag(self, part, group):
        try:
            tag = self.tags[part - 1][group - 1]
        except IndexError:
            tag = None
        return tag


class Reader:
    """ Общий фасад различных ридеров.
        Ридер должен позволять освобождать ресурсы, а также возвращать:
          - "поток" из файла источника
          - "листы" потока данных в виде словаря
          - итератор строк данных выбранного "листа"
    """
    def __init__(self):
        self._sheet = None
        self._sheet_stream = None

        self.context = None
        self.library = None
        self.filepath = None
        self.stream = None
        self.readonly = False

    def _get_library(self, producer):
        load_extra = self.context['loadExtra']
        low_memory_mode = self.context['lowMemoryMode']
        return get_entity_reader(producer,
                                 load_extra=load_extra,
                                 low_memory_mode=low_memory_mode)

    @property
    def sheet(self):
        """ Индекс листа """
        return self._sheet

    @property
    def sheets(self):
        return self.library.sheets

    @property
    def sheet_stream(self):
        return self._sheet_stream

    @property
    def sheet_names(self):
        return self.sheets.keys()

    @property
    def sheet_str(self):
        if self.context['sheetName']:
            try:
                sheet_str = self.sheets[self.sheet].name
            except KeyError:
                sheet_str = self.sheet
        else:
            sheet_str = str(self.sheet + 1)

        return sheet_str

    @property
    def sheet_identity(self):
        if self.context['sheetName']:
            return 'ИМЯ'
        else:
            return 'НОМЕР'

    @property
    def sheet_identity_type(self):
        if self.context['sheetName'] and self.context['reInSheetNames']:
            return 'МАСКА'
        return self.sheet_identity

    @property
    def sheet_hidden(self):
        return str(self.sheet_stream.sheet_state == 'hidden') if self.context['loadExtra'] else None

    @property
    def parameters(self):
        return self.library.parameters

    def activate(self, producer, context):
        self.context = context
        self.library = self._get_library(producer)

    def open_stream(self, filepath):
        self.filepath = filepath
        self.stream = self.library.get_stream(self.filepath,
                                              skip_hidden_sheets=self.context['skipHiddenSheets'],
                                              skip_hidden_row_and_column=self.context['skipHiddenParts'],
                                              detect_merged_cells=self.context['detectMergedCells'],
                                              spread_merged_rows=self.context['spreadMergedRows'],
                                              spread_merged_columns=self.context['spreadMergedColumns'],
                                              encoding=self.context['encoding'],
                                              delimiter=self.context['delimiter'],
                                              auto_detect_float=self.context['detectFloat'],
                                              auto_detect_int=self.context['detectInt'],
                                              auto_detect_datetime=self.context['detectDatetime'],
                                              path_delimiter=self.context['pathDelimiter'],
                                              dc_batch_size=self.context['dcBatchSize'])

    def open_sheet(self, sheet):
        if not self.sheet:
            self._sheet_stream = self.stream[sheet]
            self._sheet = sheet

    def close_sheet(self):
        self._sheet_stream = None
        self._sheet = None

    def get_rows(self):
        if self.context['loadExtra']:
            if self.context['readonlyExtra']:
                self.readonly = True
            log.debug(f'Расширенная загрузка. {self.sheet_identity} листа: {self.sheet_str}')

            rows = self.library.get_adapted_rows(filepath=self.filepath, sheet=self.sheet, read_only=self.readonly)
            self._sheet_stream = self.library.sheet_stream

            log.debug(f'Книга excel из файла {self.filepath} успешно открыта')
        else:
            log.debug(f'Стандартная загрузка. {self.sheet_identity} листа: {self.sheet_str}')

            rows = self.library.get_adapted_rows(sheet=self.sheet)

            log.debug(f'Лист {self.sheet_str} успешно открыт')

        return rows

    def free_resources(self):
        if self.library:
            self.library.free_resources()


class Extractor:
    """ Загружает данные из файла Excel в БД. """

    def __init__(self, source_id, run_id, producer, consumer, conn_string=None, schema=None, op_id=None,
                 op_ins_id=None):
        """ Ожидает идентификатор загрузки, строку подключения к БД и сущности в формате CDM. """
        self._filepaths = None
        self.run_id = run_id
        self.source_id = source_id
        self.op_id = op_id
        self.op_ins_id = op_ins_id
        self.producer = producer
        self.consumer = consumer
        self.ctx = defaultdict(lambda: None)
        # обработка безграмотности
        delimiter = self.producer.get_note('reInPathDelimiter') or self.producer.get_note('reInPathDelimeter')
        self.ctx['reInPathDelimiter'] = delimiter
        batch_size = self.get_batch_size()
        self.uploader = get_uploader(consumer, batch_size=batch_size, conn_string=conn_string, schema=schema)

    @property
    def filepaths(self):
        """ Возвращает список объектов Filepath, созданных из
        путей, представленных регулярными выражениями.
        """
        if self.ctx['reInPathDelimiter']:
            if self._filepaths is None:
                # под регулярное выражение могут попасть несколько файлов (self.ctx['reInPathDelimiter'] is defined)
                # и набор этих файлов еще не определялся (self._filepaths is empty)

                # находим все регулярные выражения в пути
                raw_reg_exps_re = f"{self.ctx['reInPathDelimiter']}.+?{self.ctx['reInPathDelimiter']}"
                raw_reg_exps = re.findall(raw_reg_exps_re, self.producer.path)
                # запоминаем и заменяем найденные регулярки так, чтобы они не ломали путь
                reg_exps = dict()
                for raw_reg_exp in raw_reg_exps:
                    raw_reg_exp_hash = md5(raw_reg_exp.encode(encoding='utf-8')).hexdigest()
                    self.producer.path = self.producer.path.replace(raw_reg_exp, raw_reg_exp_hash, 1)
                    reg_exps[raw_reg_exp_hash] = raw_reg_exp

                drive, path = os.path.splitdrive(self.producer.path)
                tokens = list()

                while True:
                    path, tail = os.path.split(path)
                    if tail:
                        tokens.append(tail)
                    else:
                        if path:
                            tokens.append(path)
                        break

                self._filepaths = [Filepath(drive)]
                while True:
                    new_filepaths = list()
                    token = tokens.pop()
                    len_tokens = len(tokens)

                    if token in reg_exps:
                        # обрабатываем часть пути как регулярное выражение
                        token = reg_exps[token].strip(self.ctx['reInPathDelimiter'])
                        # находим все подходящие объекты
                        for filepath in self._filepaths:
                            if os.path.isdir(filepath.path):
                                for entry in os.scandir(filepath.path):
                                    match = None
                                    if len_tokens == 0:
                                        # последняя часть пути, т.е. файлы
                                        if entry.is_file():
                                            match = re.match(token, entry.name)
                                    else:
                                        # промежуточная часть пути, т.е. папки
                                        if entry.is_dir():
                                            match = re.match(token, entry.name)
                                    if match:
                                        new_filepath = Filepath(entry.path, filepath.tags)
                                        new_filepath.append_tags(match.groups())
                                        new_filepaths.append(new_filepath)
                    else:
                        for filepath in self._filepaths:
                            new_filepath = Filepath(filepath.path, filepath.tags)
                            new_filepath.append_path(token)
                            new_filepaths.append(new_filepath)

                    self._filepaths = new_filepaths[:]

                    if len_tokens == 0:
                        break
                log.debug(f'Найдено файлов по регулярному выражению: {len(self._filepaths)}.')

        else:
            self._filepaths = [Filepath(self.producer.path)]

        return self._filepaths

    def file_not_found_handler(self, filepath: Filepath, phase: str) -> int:
        msg = f'Источник {filepath} не найден, стадия: {phase}'
        if self.ctx['ignoreAbsentFiles']:
            log.warning(msg)
            skip = -1
        else:
            raise DFExcelExtractorException(msg)

        return skip

    def file_bogus_handler(self, msg: str) -> None:
        if self.ctx['ignoreBogusFiles']:
            log.warning(msg)
        else:
            raise DFExcelExtractorException(msg)

    @staticmethod
    def get_unique_filepath(filepath: Filepath, extension=None):
        """ Генерирует уникальный путь для нового, локального, файла, с опциональной заменой расширения. """
        extension = extension or filepath.extension
        unique_base = f'{uuid4().hex}_{filepath.basename}{extension}'
        unique_path = os.path.join('/tmp', unique_base)
        return unique_path

    def get_data_tag(self, tag_name, filepath):
        """ tag_name - data tag base
            filepath - self.filepaths item
        """
        data_tag_variable = tag_name + 'Variable'
        data_tag_group = tag_name + 'Group'
        result = self.producer.get_note(data_tag_variable)

        if not result and self.ctx['reInPathDelimiter']:
            data_tag_location = self.producer.get_note(data_tag_group)
            if data_tag_location:
                try:
                    part, group = [int(i) for i in data_tag_location.split(':')]
                except ValueError:
                    msg = f'Тег {data_tag_group} содержит некорректные координаты загрузки: {data_tag_location}.'
                    msg += '\nИспользуйте формат: int:int'
                    raise DFExcelExtractorException(msg)
                result = filepath.get_tag(part, group)

        return result

    def get_extras(self):
        """ Ищет и парсит значение аннотации типа ExtraToLoad.
            Из `font.b:font.i:alignment.indent:fill.bgColor`
            создает словарь вида {[(), (), ()]}
        """
        extras = defaultdict(lambda: list())
        # определяем атрибуты, требующие дополнительных свойств и составляем их набор
        for attribute in self.producer.attributes:
            extra_to_load = attribute.get_note('extraToLoad')

            if extra_to_load:
                try:
                    tokens = extra_to_load.split(':')
                except Exception:
                    raise DFExcelExtractorException(f'extraToLoad {extra_to_load} в {attribute.name} не валидно.')

                paths = list()
                for token in tokens:
                    try:
                        path = tuple(token.split('.'))
                    except Exception:
                        raise DFExcelExtractorException(f'extraToLoad токен {path} в {attribute.name} не валидно.')
                    paths.append(path)
                extras[attribute.name] = paths
        return extras

    def get_base_folder(self):
        paths = [filepath.path for filepath in self.filepaths]
        if paths:
            base_folder = os.path.commonpath(paths)
            if os.path.isfile(base_folder):
                base_folder = os.path.dirname(base_folder)
        else:
            base_folder = None
        return base_folder

    def escape_path(self, path):
        """ Убирает двоеточия и, при необходимости, прочие невалидные символы.
            Двоеточия в именах файлов невалидны только в Windows, убирает всегда ради единообразия
        """
        decoder = {':': '='}
        return ''.join(decoder.get(c, c) for c in path)

    def archive_file(self, file_path, target_folder, base_folder, ignore_errors=False):
        """ Переносит файл в указанную архивную директорию, сохраняя структуру.
            Возвращает путь заархивированного файла, если архивация прошла успешно, иначе None.
            Если в архиве уже уже есть такой же файл, он заменяется.
        """
        new_file_path = None
        if target_folder and len(self.filepaths):
            if not base_folder:
                msg = 'Запрошена архивация, но произошла ошибка определения базового пути загрузки'
                raise DFExcelExtractorException(msg)
            relative_path = os.path.relpath(file_path, start=base_folder)
            raw_archive_path = os.path.join(target_folder, relative_path)
            archive_path = self.escape_path(raw_archive_path)
            try:
                log.info(f'Перемещаем {file_path} в архив {archive_path}')
                os.makedirs(os.path.dirname(archive_path), exist_ok=True)
                self.safely_move(file_path, archive_path)
                new_file_path = archive_path
            except Exception as e:
                msg = f'Ошибка при архивации файла {file_path} в {archive_path}: {e}'
                if ignore_errors:
                    log.warning(msg)
                else:
                    raise DFExcelExtractorException(msg)
        return new_file_path

    @staticmethod
    def safely_move(src, dst):
        """ Перемещение файла с обработкой ошибки Invalid cross-device link
        """
        try:
            return shutil.move(src, dst)
        except OSError as e:
            log.warning(f'Ошибка при перемещении файла {e}')
            shutil.copy(src, dst)
            os.remove(src)
            return os.path.join(dst, os.path.basename(src))

    def get_final_const_value(self, attribute, constant_value, is_note=False):
        """ Вытаскивает из константы конечное значение, использую заданную регулярку.
            При is_note = True, считается, что уже получено значение атрибута.
        """
        result = constant_value

        if isinstance(constant_value, str):
            constant_mask = attribute if is_note else attribute.get_note('valueMask')
            if constant_mask:
                constant_match = re.search(constant_mask, constant_value)
                if constant_match:
                    if constant_match.groups():
                        # группа 1 определена
                        result = constant_match.group(1)
                    else:
                        result = constant_match.group(0)

        return result

    def is_match(self, values, cell_re=None, row_re=None, start=1):
        """ Возвращает индекс элемента values, подходящий под регулярку cell_re, иначе - row_re.
            Для row_re предварительно склеивает значения, с индексом результата = 1.
        """
        if cell_re:
            for idx, value in enumerate(values, start=start):
                if cell_re.search(value or ''):
                    return idx
        elif row_re and row_re.search(''.join((value for value in values if value))):
            return 1

        return 0

    def get_re(self, annotation, node=None):
        """ Возвращает скомпиленную регулярку из аннотации или None. """
        node = node or self.producer
        regexp = node.get_note(annotation)
        if regexp:
            regexp = re.compile(regexp)
        return regexp

    def get_cell_range(self, cell, sheet):
        """ Возвращает координаты объединения, в которое входит ячейка. """
        result = None

        if sheet.merged_cells and isinstance(cell, MergedCell):  # @UndefinedVariable
            for r in sheet.merged_cells.ranges:
                if (r.min_col <= cell.column <= r.max_col) and (r.min_row <= cell.row <= r.max_row):
                    result = r.coord
                    break

        return result

    def get_batch_size(self):
        # Вычисляем значение rowsBatchSize на основе параметра batchSizeBreadth
        rows_batch_size = int(self.producer.get_note('rowsBatchSize', default=10000))
        batch_size_breadth = int(self.producer.get_note('batchSizeBreadth', default=15))

        if batch_size_breadth < len(self.producer.attributes):
            batch_size = batch_size_breadth * rows_batch_size
            new_rows_batch_size = batch_size // len(self.producer.attributes)

            if new_rows_batch_size < 1:
                new_rows_batch_size = 1
            rows_batch_size = new_rows_batch_size

        return rows_batch_size

    def load_data(self, rid_column, consumer_table, names_decoded, producer_columns,
                  const_fields, param_fields, extra_table, extras, head):
        """ Загружает данные в БД пачками, либо только значения, либо с дополнительными свойствами.
            TODO: инкапсулировать атомарные операции
            TODO: вынести метод в ридер
        """
        break_signal = False  # отложенная остановка для загрузки строки-хвоста
        batch = list() # пачка одновременно записываемых в базу данных
        total = 0 # количество строк загруженных данных листа
        values_dict = dict() # данные текущей строки
        blanks_counter = 0 # счетчик пустых строк, если нужен
        extra_batch = list() # пачка одновременно записываемых в базу дополнительных свойств
        extras_values_list = list() # дополнительные свойства текущей строки
        head_values_dict = dict() # данные шапки для текущей строки
        abort_signal = False # данные не подлежат загрузке, но, возможно, генерировать исключение не требуется
        row_counter = 0 # счетчик обработанных строк
        row_hidden = None # скрыта текущая строка или нет
        row_outline = None # уровень группировки текущей строки
        columns_hidden = list() # состояние скрытости столбцов строки
        columns_outline = list() # состояние группировки столбцов строки
        column_names_row = self.ctx['columnNamesRow'] # сброс номера строки с наименованиями колонок
        # сброс масок последней строки шапки
        head_last_row_mask = self.ctx['headLastRowMask']
        head_last_row_mask_l = self.ctx['headLastRowMaskL']
        # shortcuts
        load_extra = self.ctx['loadExtra']
        lazy_fields = list(const_fields.keys()) + list(param_fields.keys())
        # сброс номера первой строки шапки
        if self.ctx['minHeadRowMask'] or self.ctx['minHeadRowMaskL']:
            min_head_row = 0
        else:
            min_head_row = self.ctx['minHeadRow']

        for field in head:
            # сброс шапки предыдущего файла
            head[field]['data'] = list()
            # сброс метаданных предыдущего файла
            head[field].pop('constant', None)
            head[field]['meta']['offset'] = 0
            head[field]['chain'] = dict()

        rows = self.reader.get_rows()

        for i, r in enumerate(rows, start=1):
            row_counter += 1 # TODO: использовать i

            if abort_signal:
                # получен сигнал прерывания загрузки
                break

            if self.ctx['rowsLoadLimit'] and i > self.ctx['rowsLoadLimit']:
                # превышено максимальное количество разрешенных для обработки строк
                log.debug(f'Завершение загрузки на строке {i}. (rowsLoadLimit)')
                break

            if break_signal:
                # сигнал завершения обработки был получен на предыдущей строке
                break

            guid = None # идентификатор строки с данными для привязки к ней дополнительных свойств
            r_values = list() # список значений ячеек (только значения поддерживаются во всех библиотеках)

            # унифицируем представление значений строки
            if load_extra:
                extra_r = list() # список ячеек, не содержащий скрытые колонки, необходим для получения верных свойств

                if not self.reader.readonly:
                    # при необходимости выполняем пропуск скрытых строк или запоминаем их состояние
                    hidden = self.ctx['skipHiddenParts']
                    # запоминаем состояние строки для возможного сохранения в базу ниже
                    row_dimension = self.reader.sheet_stream.row_dimensions[i]
                    row_hidden = row_dimension.hidden is True

                    if (hidden or self.ctx['skipHiddenRows']) and row_hidden:
                        # пропускаем скрытые строки
                        continue
                    else:
                        # преобразуем в строку, чтобы в таблицы сохранилось строковое True или False
                        row_hidden = str(row_hidden)

                    row_outline = row_dimension.outlineLevel

                for idx, cell in enumerate(r, start=1):
                    extra_r.append(cell)

                    # TODO #137 2pymancer: нужна ли тут какая-то обработка в рамках задачи columnLetter?
                    if not self.reader.readonly:
                        # при необходимости выполняем пропуск скрытых столбцов или запоминаем их состояние
                        letter = get_column_letter(idx)
                        # запоминаем состояние столбца для возможного сохранения в базу ниже
                        column_dimension = self.reader.sheet_stream.column_dimensions[letter]
                        hidden_col = column_dimension.hidden is True

                        if (hidden or self.ctx['skipHiddenColumns']) and hidden_col:
                            # пропускаем скрытые колонки
                            extra_r.pop()
                            continue
                        else:
                            columns_hidden.append(str(hidden_col))

                        columns_outline.append(column_dimension.outlineLevel)

                    cell_value = None if cell.value is None or cell.value == '' else str(cell.value)
                    r_values.append(cell_value)
            else:
                for value in r:
                    value = None if value is None or value == '' else str(value)
                    r_values.append(value)

            if head and not min_head_row and self.is_match(r_values,
                                                           cell_re=self.ctx['minHeadRowMask'],
                                                           row_re=self.ctx['minHeadRowMaskL']):
                # первая строка шапки определяется маской
                min_head_row = i

            if head and (min_head_row and i >= min_head_row) and (i <= min_head_row + self.ctx['maxHeadRowOffset']):
                # первая строка шапки определена и текущая строка в диапазоне шапки
                for field in head:
                    if not head[field].get('constant'):
                        # значение атрибута не определено, запоминаем сложную шапку, считая, что она с первой строки
                        first_row_known = True
                        if head[field]['meta']['first'] and not head[field]['meta']['offset']:
                            # первая строка шапки смещена
                            first_row_known = False
                            if self.is_match(r_values, cell_re=head[field]['meta']['first']):
                                # действительная первая строка шапки определена
                                first_row_known = True
                                head[field]['meta']['offset'] = i - 1

                        head_offset = head[field]['meta']['offset']
                        ordinal_with_offset = head[field]['meta']['ordinal'] + head_offset
                        chain_mask = head[field]['meta']['chainMask']

                        if chain_mask and (not head_offset or i < ordinal_with_offset):
                            # нашли первую строку шапки для данного атрибута, собираем цепочку значений
                            for idx, value in enumerate(r_values, start=1):
                                head[field]['chain'][idx] = f"{head[field]['chain'].get(idx, '')}{value or ''}"

                        if ordinal_with_offset == i and first_row_known:
                            constant_idx = None
                            constant_mask = head[field]['meta']['columnMask']

                            if constant_mask:
                                # номер колонки константы определяется по маске
                                constant_idx = self.is_match(r_values, cell_re=constant_mask)

                            constant_idx = constant_idx or head[field]['meta']['column']

                            # TODO #137 2pymancer: нужна ли тут какая-то обработка в рамках задачи columnLetter?
                            if chain_mask:
                                # номер колонки определяется по маске, с учетом ячеек выше до строки с 'firstRowMask'
                                for idx, value in enumerate(r_values, start=1):
                                    if chain_mask.search(f"{head[field]['chain'].get(idx, '')}{value or ''}"):
                                        constant_idx = idx
                                        break

                            if constant_idx:
                                # одно значение распространяется на всю строку
                                head[field]['constant'] = None

                                try:
                                    # нумерация столбцов с 1 для пользователя, но с 0 для библиотеки
                                    constant = r_values[constant_idx - 1]
                                except IndexError:
                                    constant = None

                                if constant:
                                    if head[field]['meta']['columnOrdinal']:
                                        # номер колонки этой строки так же является номером колонки значения показателя
                                        producer_columns[head[field]['meta']['target']] = constant_idx

                                    # записываем значение константы из ячейки
                                    constant_mask = head[field]['meta']['mask']
                                    head[field]['constant'] = self.get_final_const_value(constant_mask,
                                                                                         constant,
                                                                                         is_note=True)
                            else:
                                # значение зависит от столбца
                                for value in r_values:
                                    head[field]['data'].append(value)

            # пропускаем строки до строки с наименованиями
            if not names_decoded:
                if i < column_names_row:
                    # еще не дошли до строки с наименованиями колонок
                    continue
                elif self.ctx['columnNamesRowMask'] or self.ctx['columnNamesRowMaskL'] or (i == column_names_row):

                    if self.ctx['columnNamesRowMask'] or self.ctx['columnNamesRowMaskL']:
                        # определяем номер строки с наименованиями колонок
                        if self.is_match(r_values,
                                         cell_re=self.ctx['columnNamesRowMask'],
                                         row_re=self.ctx['columnNamesRowMaskL']):
                            column_names_row = i

                    if i == column_names_row:
                        # сопоставляем наименования колонок и атрибутов
                        names_decoded = True
                        decoded = list()

                        for idx, value in enumerate(r_values, start=1):
                            row_value = value.strip() if value else ''

                            for attribute in self.producer.attributes:
                                column_name = attribute.columnName or attribute.name

                                if self.ctx['reInColumnNames']:
                                    matched_column_name = re.search(column_name, row_value)
                                else:
                                    matched_column_name = column_name == row_value

                                if matched_column_name:
                                    if attribute.name in decoded:
                                        # наименования колонок в строке не уникальны
                                        msg = 'Коллизия колонки <{}> ({}) и атрибута <{}>.'
                                        msg = msg.format(column_name, producer_columns[attribute.name], attribute.name)
                                        if self.ctx['reInColumnNames']:
                                            msg += ' Включен поиск колонок по регулярным выражениям.'
                                        raise DFExcelExtractorException(msg)
                                    else:
                                        if not producer_columns[attribute.name]:
                                            # задаем колонку если она не пришла из сложной шапки и нет коллизии в файле
                                            producer_columns[attribute.name] = idx
                                            decoded.append(attribute.name)
                        # строку с наименованиями грузить не нужно
                        continue
                    elif self.ctx['columnNamesRow'] and i > self.ctx['columnNamesRow']:
                        msg = f'Строка {self.ctx["columnNamesRow"]} с именами колонок пропущена'
                        raise DFExcelExtractorException(msg)
                    else:
                        # строка с наименованиями еще не определена либо не достигнута
                        continue
                else:
                    raise DFExcelExtractorException('Не удалось сопоставить колонки и атрибуты')

            if i <= self.ctx['rowsNumberToSkip']:
                # еще не дошли до строк, которые нужно загружать
                continue

            if ((self.ctx['rowMaskToSkip'] or self.ctx['rowMaskToSkipL'])
                and self.is_match(r_values, cell_re=self.ctx['rowMaskToSkip'], row_re=self.ctx['rowMaskToSkipL'])):
                # если строка соответствует переданному регулярному выражению - пропускаем
                continue

            # пропускаем шапку
            if head_last_row_mask or head_last_row_mask_l:
                if column_names_row and column_names_row < i:
                    # шапка должна быть до строки с наименованиями
                    head_last_row_mask = head_last_row_mask_l = None
                else:
                    if self.ctx['headLastRowMaskStrict']:
                        # оставляем только значения из загружаемых колонок
                        head_values = (hv for iv, hv in enumerate(r_values) if iv in producer_columns.values())
                    else:
                        head_values = r_values

                    if self.is_match(head_values, cell_re=head_last_row_mask, row_re=head_last_row_mask_l):
                        head_last_row_mask = head_last_row_mask_l = None
                        if self.ctx['skipHeadLastRow']:
                            continue
                    else:
                        continue

            if ((self.ctx['tailFirstRowMask'] or self.ctx['tailFirstRowMaskL'])
                and bool(self.is_match(r_values,
                                       cell_re=self.ctx['tailFirstRowMask'],
                                       row_re=self.ctx['tailFirstRowMaskL']))):
                # пропускаем подвал
                if self.ctx['skipTailFirstRow']:
                    log.debug(f'Завершение загрузки на строке {i}. (tailFirstRowMask)')
                    break
                else:
                    log.debug(f'Завершение загрузки на строке {i+1}. (tailFirstRowMask + skipTailFirstRow)')
                    break_signal = True

            if self.ctx['multipleColumnsAttribute']:
                # загружаем все колонки последовательно в один атрибут
                cur_rows = [(cr, ) for cr in r_values]
            elif self.ctx['pivotAttribute']:
                # нумерация столбцов в producer_columns с 1 для пользователя, но с 0 для библиотеки
                pivot_idx = producer_columns[self.ctx['pivotAttribute']] - 1
                pivot_heads_values = list()
                blanks_columns = list()
                # нумерация с 1 и при этом нам нужен следующий столбец за данными помеченными как isPivot
                k = pivot_idx + 2
                # Поддержка нескольких атрибутов в шапке
                for head_attribute, head_param in head.items():
                    pivot_heads_values.append(head_param['data'])
                    blanks_columns.append(head_param['meta']['maxBlanks'])
                    producer_columns[head_attribute] = k
                    k += 1
                cur_rows = self.split_row_for_pivot(r_values, pivot_idx, pivot_heads_values, blanks_columns)
            else:
                # загружаем всю строку сразу
                cur_rows = [r_values]

            for cur_idx, cur_row in enumerate(cur_rows):
                # собираем значения текущей строки
                values_dict.clear()
                if load_extra:
                    # инициализация сохранения дополнительных свойств
                    del extras_values_list[:]
                    guid = str(uuid4())

                for k in producer_columns:
                    producer_column = producer_columns[k]
                    if not producer_column:
                        if k in lazy_fields:
                            # значение уже известно и будет проставлено позже
                            continue
                        if self.ctx['nullMappingErrors']:
                            # включено зануление отсутствующих в файле столбцов
                            values_dict[k] = None
                            continue
                        msg = f'Столбца для <{k}> нет в <{self.reader.filepath}>, метаданные не соответствуют файлу?'
                        if self.ctx['skipMappingErrors']:
                            msg = f'\n{msg}. Загрузка файла прервана.'
                            log.warning(msg)
                            abort_signal = True
                            break
                        else:
                            raise DFExcelExtractorException(msg)
                    if len(cur_row) >= producer_column:
                        # нумерация столбцов в producer_columns с 1 для пользователя, но с 0 для библиотеки
                        producer_column_index = producer_column - 1

                        try:
                            value = cur_row[producer_column_index]
                        except IndexError:
                            msg = (f'Столбца {producer_column_index} нет в <{self.reader.filepath}>,'
                                   f' нумерация в метаданных <{self.consumer.name}> не с 1?')
                            raise DFExcelExtractorException(msg)

                        values_dict[k] = value

                        if extras and k in extras:
                            # сохраняем дополнительные свойства атрибутов, которым это требуется
                            # учитываем, что может грузиться только текущая колонка
                            cell_idx = producer_column_index + cur_idx
                            cell = extra_r[cell_idx] # используется extra_r, так в нем колонки соответствуют r_values
                            ev = extras[k]

                            for ek in ev:
                                nroe = not self.ctx['readonlyExtra']
                                # определяем специальные расширенные свойства
                                if nroe and ek == ('range', ):
                                    # принадлежность ячейки к объединению (рассчетное свойство)
                                    extra_value = self.get_cell_range(cell, self.reader.sheet_stream)
                                elif nroe and self.reader.sheet_hidden is not None and ek == ('sheet', 'hidden'):
                                    extra_value = self.reader.sheet_hidden
                                elif row_hidden is not None and ek == ('row', 'hidden'):
                                    extra_value = row_hidden
                                elif columns_hidden and ek == ('column', 'hidden'):
                                    extra_value = columns_hidden[cell_idx]
                                elif row_outline is not None and ek == ('row', 'outline'):
                                    extra_value = row_outline
                                elif columns_outline and ek == ('column', 'outline'):
                                    extra_value = columns_outline[cell_idx]
                                else:
                                    # определяем расширенные свойства
                                    extra_value = get_prop_by_path(cell, ek)

                                # запоминаем расширенные свойства
                                extras_values_list.append(
                                    {'row_id': guid,
                                     'column_name': k,
                                     'key': '.'.join(ek),
                                     'value': extra_value
                                     }
                                )

                        if head:
                            # добавляем шапку, если есть
                            for field in head:
                                field_dict = head[field]
                                field_meta = field_dict['meta']

                                if 'constant' in field_dict:
                                    # записываем константу в соответствующий атрибут строки
                                    head_values_dict[field] = field_dict['constant']
                                elif (len(field_dict['data']) > producer_column_index
                                    and not field_meta['skip'] and field_meta['target'] and field_meta['target'] == k):
                                        head_values_dict[field] = field_dict['data'][producer_column_index]

                    else:
                        values_dict[k] = None

                if abort_signal:
                    # некачественные данные, пропускаем файл
                    break
                else:
                    if self.ctx['skipBlankRows'] and not any(v is not None for k, v in values_dict.items()):
                        # пустую строку не сохраняем, если она не принадлежит объединению в расширенной загрузке
                        merged_cell = False

                        if load_extra and self.ctx['detectMergedCells']:
                            for v in extras_values_list:
                                if v['key'] == 'range' and v['value'] is not None:
                                    merged_cell = True
                                    break

                        if not merged_cell:
                            if self.ctx['maxBlanksToProcess'] is not None:
                                blanks_counter += 1
                                if blanks_counter > self.ctx['maxBlanksToProcess']:
                                    # похоже, что данных больше нет, завершаем обработку
                                    log.debug(f'Завершение загрузки на строке {i}. (maxBlanksToProcess)')
                                    break_signal = True
                                    break

                            continue

                    # заполняем параметры, если значения не пустые
                    values_dict.update(param_fields)
                    # заполняем константы, если значения не пустые
                    values_dict.update(const_fields)
                    # добавляем значения служебных полей, если значения не пустые
                    if guid:
                        # добавляем guid только сейчас, чтобы не мешать определению пустых строк
                        values_dict[rid_column.name] = guid
                    if head:
                        # добавляем данные шапки, если есть
                        values_dict.update(head_values_dict)
                    # добавляем строку в пачку для вставки
                    batch.append(values_dict.copy())
                    if extras_values_list:
                        # добавляем дополнительные свойства в пачку для вставки, убрав отсутствующие в данных
                        extra_batch.extend(extras_values_list)

                    if self.ctx['rowsBatchSize'] and i % self.ctx['rowsBatchSize'] == 0:
                        # собрали пачку для разовой, одномоментной, вставки
                        log.debug(f'Получено строк: {i}')
                        self.uploader.upload(batch, container=consumer_table)
                        log.debug(f'Загружено строк: {i}')
                        total += len(batch)
                        del batch[:]
                        # сохраняем соответствующее форматирование
                        if extra_batch:
                            self.uploader.upload(extra_batch, container=extra_table)

                            log.debug('Загружены метаданные строк')
                            del extra_batch[:]

        if batch:
            # сохраняем остатки, не составившие полную пачку
            log.debug(f'Получено {i} строк')
            self.uploader.upload(batch, container=consumer_table)
            log.debug(f'Загружено {i} строк')
            total += len(batch)
            del batch[:]

            if extra_batch:
                # сохраняем остатки дополнительных свойств, не составившие полную пачку
                self.uploader.upload(extra_batch, container=extra_table)

                log.debug('Загружены метаданные строк')
                del extra_batch[:]

        self.uploader.commit()
        log.debug(f'Обработано строк: {row_counter}.')

        return total

    @staticmethod
    def find_blanks_index(values, max_blank_count, start=0):
        """ Возвращает индекс начала первого вхождения последовательности, состоящей из более чем n пустых значений
        :param values: список проверяемых значений
        :param max_blank_count: допустимое количество пустых значений
        :param start: позиция начала поиска
        :return:
        """
        j = -1
        if max_blank_count is not None:
            blanks_counter = 0
            for i in range(len(values)):
                if i < start:
                    continue
                value = values[i]
                if value is None or value == '':
                    blanks_counter += 1
                else:
                    blanks_counter = 0
                if blanks_counter > max_blank_count:
                    j = i - max_blank_count
                    break
        return j

    def split_row_for_pivot(self, row, pivot_idx, col_heads_values, blanks_columns=None):
        """ Разбивает список на две части: постоянную (до индекса) и переменную (после). Постоянную часть (боковик и
        шапку) и каждый элемент переменной части объединяет в новые списки и возвращает результат.
        Используется для формирования строк данных при загрузке кросс-таблиц
        :param row: строка с данными
        :param pivot_idx: индекс элемента начала переменной части (нумерация с 0)
        :param col_heads_values: список шапок с данными (добавляется после переменной части справа)
        :param blanks_columns: допустимые максимальное количество пустых столбцов для шапок (используется для
        ограничения длины строки)
        :return:
        """

        # TODO 70: выполнять только один раз
        max_col_head_len = 0
        max_col_head_blank = 0
        for i, col_head_values in enumerate(col_heads_values):
            blank_columns = blanks_columns[i]
            # Обрезаем шапку до первых пустых колонок, число которых больше заданного
            blanks_index = self.find_blanks_index(col_head_values, blank_columns, start=pivot_idx)
            if blanks_index > 0:
                col_head_values = col_head_values[0:blanks_index]
            if len(col_head_values) > max_col_head_len:
                max_col_head_len = len(col_head_values)
                max_col_head_blank = blank_columns

        # Обрезание или наращивание строки
        n = max_col_head_len - len(row)
        if n > 0:
            # Если в строке есть пустой "хвост", то дополняем ее пустыми значениями до длины шапки
            row += [None for _ in range(n)]
        else:
            # Если строка больше самой длинной шапки - обрежем строку до нее с учетом допустимых пустых значения в шапке
            if len(row) > max_col_head_len + max_col_head_blank:
                row = row[0:max_col_head_len + max_col_head_blank]

        # Если длины шапки недостаточно для строки, дополняем ее пустыми значениями до длины строки, но не более,
        # чем длина шапки + blank_columns
        max_head_tail = 0  # максимально допустимое наращивание хвоста шапки пустыми значениями
        for i, col_head_values in enumerate(col_heads_values):
            n = len(row) - len(col_head_values)
            if n > 0:
                head_tail = min(n, blanks_columns[i])
                if head_tail > max_head_tail:
                    max_head_tail = head_tail  # ?
                col_head_values += [None for _ in range(head_tail)]

        rows = list()
        row_head_values = row[:len(row) - len(row[pivot_idx:])]
        pivot_values = row[pivot_idx:]

        for i, value in enumerate(pivot_values):
            cur_row = list()
            cur_row += row_head_values
            cur_row.append(value)
            for col_head_values in col_heads_values:
                cur_row.append(col_head_values[pivot_idx + i])
            rows.append(cur_row)
        return rows

    def run(self):
        """ Загрузка данных из файла в базу.
            TODO: при наличии нескольких разных масок, по возможности проверять их в одном итераторе строки
            TODO оптимизировать получение аннотаций через get_notes
        """
        processed_files = dict() # обработанные файлы, True - файл успешно обработан, False - нет
        # значения аннотации `const` атрибутов источника, требующие особой обработки
        special_consts = ['sheetName', 'fileName', 'filePath']

        self.ctx['rowsLoadLimit'] = int(self.producer.get_note('rowsLoadLimit', default=0))

        self.ctx['skipBlankRows'] = self.producer.get_note('skipBlankRows', default=True)

        max_blanks = self.producer.get_note('maxBlanksToProcess')
        self.ctx['maxBlanksToProcess'] = None if max_blanks is None else int(max_blanks)

        # у PIPwareEntity первая строка - наименования колонок
        default_skip = 1 if type(self.producer) in (PIPwareWebEntity, DCDimensionEntity) else 0
        self.ctx['rowsNumberToSkip'] = int(self.producer.get_note('rowsNumberToSkip', default=default_skip))

        self.ctx['columnNamesRow'] = int(self.producer.get_note('columnNamesRow', default=default_skip))
        self.ctx['columnNamesRowMask'] = self.get_re('columnNamesRowMask')
        if not self.ctx['columnNamesRowMask']:
            self.ctx['columnNamesRowMaskL'] = self.get_re('columnNamesRowMaskL')

        self.ctx['rowMaskToSkip'] = self.get_re('rowMaskToSkip')
        if not self.ctx['rowMaskToSkip']:
            self.ctx['rowMaskToSkipL'] = self.get_re('rowMaskToSkipL')

        self.ctx['headLastRowMask'] = self.get_re('headLastRowMask')
        if not self.ctx['headLastRowMask']:
            self.ctx['headLastRowMaskL'] = self.get_re('headLastRowMaskL')
        if self.ctx['headLastRowMask'] or self.ctx['headLastRowMaskL']:
            self.ctx['headLastRowMaskStrict'] = self.producer.get_note('headLastRowMaskStrict', default=False)
        else:
            self.ctx['headLastRowMaskStrict'] = False

        self.ctx['skipHeadLastRow'] = self.producer.get_note('skipHeadLastRow', default=True)

        self.ctx['tailFirstRowMask'] = self.get_re('tailFirstRowMask')
        if not self.ctx['tailFirstRowMask']:
            self.ctx['tailFirstRowMaskL'] = self.get_re('tailFirstRowMaskL')

        self.ctx['skipTailFirstRow'] = self.producer.get_note('skipTailFirstRow', default=True)

        self.ctx['reInColumnNames'] = self.producer.get_note('reInColumnNames', default=False)

        self.ctx['detectMergedCells'] = self.producer.get_note('detectMergedCells', default=False)

        self.ctx['skipHiddenSheets'] = self.producer.get_note('skipHiddenSheets', default=True)
        self.ctx['skipHiddenParts'] = self.producer.get_note('skipHiddenParts', default=True)
        self.ctx['skipHiddenRows'] = self.producer.get_note('skipHiddenRows', default=False)
        self.ctx['skipHiddenColumns'] = self.producer.get_note('skipHiddenColumns', default=False)

        self.ctx['reInSheetNames'] = self.producer.get_note('reInSheetNames', default=False)
        self.ctx['sheetName'] = self.producer.get_note('sheetName')
        # индексация листа начинается с 0, но для пользователя в метаданных нумерация с 1
        self.ctx['sheetOrdinal'] = abs(int(self.producer.get_note('sheetOrdinal', default=1)) - 1)

        self.ctx['skipMappingErrors'] = self.producer.get_note('skipMappingErrors', default=False)
        self.ctx['nullMappingErrors'] = self.producer.get_note('nullMappingErrors', default=False)

        self.ctx['multiAccessMode'] = self.producer.get_note('multiAccessMode', default=False)
        self.ctx['doNotPipeSkipped'] = self.producer.get_note('doNotPipeSkipped', default=False)
        self.ctx['ignoreAbsentFiles'] = self.producer.get_note('ignoreAbsentFiles', default=False)
        self.ctx['ignoreBogusFiles'] = self.producer.get_note('ignoreBogusFiles', default=False)
        self.ctx['ignoreAbsentSheets'] = self.producer.get_note('ignoreAbsentSheets', default=False)

        self.ctx['minHeadRow'] = int(self.producer.get_note('minHeadRow', default=1))
        self.ctx['maxHeadRowOffset'] = int(self.producer.get_note('maxHeadRowOffset', default=20))

        self.ctx['minHeadRowMask'] = self.get_re('minHeadRowMask')
        if not self.ctx['minHeadRowMask']:
            self.ctx['minHeadRowMaskL'] = self.get_re('minHeadRowMaskL')

        self.ctx['ignoreArchiveErrors'] = self.producer.get_note('ignoreArchiveErrors', default=False)
        self.ctx['removeArchivedFolders'] = self.producer.get_note('removeArchivedFolders', default=False)

        self.ctx['loadMultipleColumns'] = self.producer.get_note('loadMultipleColumns', default=False)
        self.ctx['loadMultipleSheets'] = self.producer.get_note('loadMultipleSheets', default=False)

        self.ctx['lowMemoryMode'] = self.producer.get_note('lowMemoryMode')

        # аннотации специфичные для NetDB
        self.ctx['spreadMergedRows'] = self.producer.get_note('spreadMergedRows', default=False)
        self.ctx['spreadMergedColumns'] = self.producer.get_note('spreadMergedColumns', default=False)

        # аннотации специфичные для CSV
        self.ctx['encoding'] = self.producer.get_note('encoding', default='utf-8')
        self.ctx['delimiter'] = self.producer.get_note('delimiter', default=',')
        self.ctx['detectFloat'] = self.producer.get_note('detectFloat', default=False)
        self.ctx['detectInt'] = self.producer.get_note('detectInt', default=False)
        self.ctx['detectDatetime'] = self.producer.get_note('detectDatetime', default=False)

        # аннотации dc dimension
        self.ctx['pathDelimiter'] = self.producer.get_note('pathDelimiter', default='/')
        self.ctx['dcBatchSize'] = int(self.producer.get_note('dcBatchSize', default=1000))

        # определяем базовую  архивную директории загружаемых файлов
        archive_folder = self.producer.get_note('archiveFolder')
        failed_folder = self.producer.get_note('failedFolder')
        base_folder = self.get_base_folder() if (archive_folder or failed_folder) else None

        # конвертировать ли не xlsx файлы
        self.ctx['convertToXLSX'] = self.producer.get_note('convertToXLSX', default=False)

        # служебные поля
        const_fields_global = dict()
        # параметры данных
        param_fields = dict()

        # получаем ридер, необходимого для загрузки целевого файла
        self.reader = Reader()

        try:
            grand_total = 0 # общее количество загруженных записей
            head = dict() # словарь хранения данных сложной шапки
            # атрибуты в соответствии с их номерами или наименованиями столбцов
            producer_columns = dict()
            # таблица для вставки
            table_name_full = self.uploader.get_container_name(full=True)

            if self.ctx['columnNamesRow'] or self.ctx['columnNamesRowMask'] or self.ctx['columnNamesRowMaskL']:
                names_decoded = False
            else:
                names_decoded = True

            for attribute in self.producer.attributes:
                ordinal = None
                attribute_is_row = attribute.get_note('isRow', default=False)

                const_value = None if attribute_is_row else attribute.get_note('const')
                param_value = None if attribute_is_row else attribute.get_note('param')

                if hasattr(attribute, 'ordinal'):
                    if getattr(attribute, 'columnLetter', None):
                        attribute.ordinal = column_index_from_string(attribute.columnLetter)

                    try:
                        ordinal = int(attribute.ordinal)
                    except TypeError:
                        pass

                if (not names_decoded) and (not attribute_is_row):
                    # присваивание здесь ordinal, а не сразу None позволяет задавать ordinal для отдельных колонок,
                    # а также автоматически выставлять 1 строку данных как строку наименований колонок
                    # для сущностей типа PIPwareWebEntity, DCDimensionEntity
                    producer_columns[attribute.name] = ordinal
                else:
                    if ordinal is None:
                        raise DFExcelExtractorException(f'Не задана колонка для атрибута <{attribute.name}>')

                    if attribute_is_row and not self.ctx['loadMultipleColumns']:
                        # атрибут записывает данные определенной строки или ячейки
                        head[attribute.name] = dict()
                        head[attribute.name]['meta'] = dict()
                        head[attribute.name]['meta']['ordinal'] = ordinal
                        head[attribute.name]['meta']['offset'] = 0

                        column_value_mask = attribute.get_note('firstRowMask')
                        head[attribute.name]['meta']['first'] = column_value_mask and re.compile(column_value_mask)

                        head_column = int(attribute.get_note('rowColumn', default=0))
                        head[attribute.name]['meta']['column'] = head_column

                        head[attribute.name]['meta']['columnMask'] = self.get_re('rowColumnMask', node=attribute)

                        head[attribute.name]['meta']['chainMask'] = self.get_re('rowChainMask', node=attribute)

                        head[attribute.name]['meta']['mask'] = self.get_re('valueMask', node=attribute)

                        target = attribute.get_note('targetAttribute')
                        head[attribute.name]['meta']['target'] = target

                        do_not_load = attribute.get_note('doNotLoad', default=False)
                        head[attribute.name]['meta']['skip'] = do_not_load

                        set_column_ordinal = attribute.get_note('setColumnOrdinal', default=False)
                        head[attribute.name]['meta']['columnOrdinal'] = set_column_ordinal if target else False

                        blank_cols = attribute.get_note('maxBlanksToProcess')
                        head[attribute.name]['meta']['maxBlanks'] = None if blank_cols is None else int(blank_cols)
                    else:
                        producer_columns[attribute.name] = ordinal

                if const_value is not None:
                    # обрабатываем маску для уже определенных значений
                    # значение None (Null) может быть записано с помощью маски, которая точно ничего не найдет
                    if const_value not in special_consts:
                        # для специальных констант регулярка применяется при их определении
                        const_value = self.get_final_const_value(attribute, const_value)
                    const_fields_global[attribute.name] = const_value

                if param_value is not None:
                    param_fields[attribute.name] = param_value

                # Признак загрузки данных для кросс-таблицы
                if attribute.get_note('isPivot', default=False):
                    if self.ctx['pivotAttribute']:
                        msg = 'Обнаружено более одного атрибута, отмеченного как начало данных кросс-таблицы'
                        raise DFExcelExtractorException(msg)
                    self.ctx['pivotAttribute'] = attribute.name

            if self.ctx['loadMultipleColumns']:
                # все колонки должны быть загружены в один атрибут, если таковой есть
                multi_columns_attributes = set(producer_columns) - set(const_fields_global)
                len_multi_columns_attributes = len(multi_columns_attributes)
                if len_multi_columns_attributes == 1:
                    # определен атрибут для загрузки всех колонок
                    names_decoded = True
                    self.ctx['multipleColumnsAttribute'] = multi_columns_attributes.pop()
                    producer_columns[self.ctx['multipleColumnsAttribute']] = 1
                elif len_multi_columns_attributes > 1:
                    msg = 'Обнаружено более одного атрибута, предназначенного для загрузки нескольких колонок'
                    raise DFExcelExtractorException(msg)

            # номер строки с наименованиями не может быть меньше обрабатываемых
            if self.ctx['columnNamesRow'] and self.ctx['columnNamesRow'] > self.ctx['rowsNumberToSkip'] + 1:
                msg = 'Данные с rowsNumberToSkip не могут начинаться раньше строки с именами колонок columnNamesRow'
                raise DFExcelExtractorException(msg)

            producer_strict_mapping = self.producer.get_note('enforceStrictMapping', default=True)
            consumer_strict_mapping = self.consumer.get_note('enforceStrictMapping', default=True)

            consumer_roles = dict()
            all_producer_columns = list(head.keys()) + list(producer_columns.keys())

            for attribute in self.consumer.attributes:
                role = attribute.get_note('role')
                # кэшируем роли
                consumer_roles[attribute.name] = role
                # контролируем соответствие атрибутов источника и получателя (без учета полей с ролями)
                enforce_producer_strict_mapping = producer_strict_mapping
                if enforce_producer_strict_mapping and (not role) and (attribute.name not in all_producer_columns):
                    msg = f'В источнике {self.producer.name} отсутствует атрибут получателя {attribute.name}'
                    raise DFExcelExtractorException(msg)

            if self.consumer.get_note('precreateEntity', default=False):
                # создаем таблицу автоматически, если ее нет
                self.uploader.create()

            if self.consumer.get_note('purgeExistingData', default=False):
                self.uploader.empty()

            consumer_table = self.uploader.get_container()

            # consumer_table не приводится к boolean, проверяем на None явно
            if consumer_table is None:
                # проверяем требуется ли данная сущность, если нет, то просто пропускаем загрузку
                if consumer_strict_mapping or producer_strict_mapping:
                    raise DFExcelExtractorException(f'Таблица {table_name_full} не найдена в БД')
            else:
                if consumer_strict_mapping:
                    differences = list()
                    # проверяем, что таблица-получатель в БД соответствует метаданным сущности-получателя
                    for attribute in self.consumer.attributes:
                        # TODO: добавить проверку на тип колонки
                        if attribute.name not in consumer_table.columns:
                            differences.append(attribute.name)

                    if differences:
                        # в описании сущности есть новые атрибуты
                        msg = f'В таблице-получателе {table_name_full} отсутствуют атрибуты:\n'
                        msg += '\n'.join(differences)

                        if self.consumer.get_note('recreateIfOutOfSync'):
                            log.info(msg)
                            self.uploader.drop()
                            consumer_table = self.uploader.create()
                            log.warning(f'Таблица-получатель {table_name_full} пересоздана.')
                        else:
                            raise DFExcelExtractorException(msg)

                # определяем загружаемый лист по наименованию или номеру
                # лист должен подходить каждому из загружаемых файлов
                sheet = None

                # определяем дополнительные свойства для загрузки, если требуется
                extras = None
                extra_table = None
                rid_column = None

                self.ctx['loadExtra'] = self.producer.get_note('loadExtra', default=False)

                if self.ctx['loadExtra']:
                    rid_column = rid_column or self.consumer.find_by_note('attributes', 'role', 'rid')

                    if not rid_column:
                        raise DFExcelExtractorException(f'В таблице {table_name_full} нет поля с ролью rid')
                    elif rid_column.dataType != DataType.Guid:
                        msg = f'Поле {table_name_full}.{rid_column.name} должно быть типа guid'
                        raise DFExcelExtractorException(msg)

                    extras = self.get_extras()
                    if not extras:
                        self.ctx['loadExtra'] = False
                        log.warning('Не определены дополнительные свойства. Расширенная загрузка отключена.')
                    else:
                        extra_table = self.uploader.create(rid_attr=rid_column.name)

                        self.ctx['readonlyExtra'] = self.producer.get_note('readonlyExtra', default=True)

                        detect_merged = self.ctx['detectMergedCells']

                        if self.ctx['readonlyExtra']:
                            msg = ''
                            if detect_merged:
                                msg = f'{msg}`detectMergedCells` = `true` '
                            if self.ctx['skipHiddenParts']:
                                msg = f'{msg}`skipHiddenParts` = `true` '
                            if self.ctx['skipHiddenRows']:
                                msg = f'{msg}`skipHiddenRows` = `true` '
                            if self.ctx['skipHiddenColumns']:
                                msg = f'{msg}`skipHiddenColumns` = `true` '
                            if self.ctx['skipHiddenSheets']:
                                msg = f'{msg}`skipHiddenSheets` = `true` '
                            if msg:
                                msg = f'При {msg}`readonlyExtra` необходимо выставить в `false`'
                                raise DFExcelExtractorException(msg)

                        if detect_merged:
                            msg = 'В расширенной загрузке диапазон объединения определяется по свойству `range`.'
                            log.debug(msg)

                self.reader.activate(producer=self.producer, context=self.ctx)

                for filepath in self.filepaths:
                    """
                    причина пропуска: -1 - нет файла, 1 - нет листа, 2 - не найден лист, 3 - неизвестная ошибка, 
                    4 - некорректный формат файла; < 0 - не пайпить
                    """
                    log.info(f'Старт загрузки данных из источника {filepath} в таблицу {table_name_full}')
                    skip = 0
                    file_total = 0  # количество загруженных записей одного файла

                    known_exts = ['.xls', 'xlsb', 'ods']
                    xlsx_ext = '.xlsx'
                    xls_ext = '.xls'
                    xlsb_ext = '.xlsb'
                    if filepath.extension in known_exts:
                        ext = filepath.extension

                        if self.ctx['lowMemoryMode']:
                            raise DFExcelExtractorException(f'для {ext} не поддерживается экономия памяти')
                        if self.ctx['loadExtra']:
                            raise DFExcelExtractorException(f'{ext} не поддерживается экстракция форматирования')

                        if self.ctx['convertToXLSX']:

                            if self.ctx['multiAccessMode']:
                                raise DFExcelExtractorException('Режим конвертации не поддерживает multiAccessMode')

                            log.warning(f'Экспериментальный функционал: экстракт из конвертированного {ext}.')
    
                            unique_path = self.get_unique_filepath(filepath, extension=xlsx_ext)
    
                            try:
                                if ext == xls_ext:
                                    pyexcel.save_book_as(file_name=filepath.path,
                                                         dest_file_name=unique_path,
                                                         skip_hidden_sheets=self.ctx['skipHiddenSheets'],
                                                         skip_hidden_row_and_column=self.ctx['skipHiddenParts'],
                                                         detect_merged_cells=self.ctx['detectMergedCells'])
                                else:
                                    pyexcel.save_book_as(file_name=filepath.path, dest_file_name=unique_path)
                            except XLRDError as e:
                                msg = f'Ошибка преобразования {filepath.path} в {xlsx_ext}. Ошибка: {e}'
                                skip = 4
                                self.file_bogus_handler(msg)
                            except FileNotFoundError:
                                skip = self.file_not_found_handler(filepath, f'конвертация из {ext} в {xlsx_ext}')
                            else:
                                filepath.path = unique_path
                        else:
                            if ext in (xlsb_ext, ):
                                if self.ctx['skipHiddenParts']:
                                    log.warning(f'для {ext} не поддерживаются скрытые ячейки')
                                if self.ctx['skipHiddenSheets']:
                                    log.warning(f'для {ext} не поддерживаются скрытые листы')

                            log.warning(f'Экспериментальный функционал: экстракт из {ext}.')

                    if not skip:
                        date_begin = self.get_data_tag('dateBegin', filepath)
                        date_end = self.get_data_tag('dateEnd', filepath)
                        unit_id = self.get_data_tag('unitId', filepath)

                        # служебные поля каждого файла, в случае отличия
                        const_fields_file = const_fields_global.copy()

                        # значения сервисных атрибутов по ролям
                        const_values = {'id': None, 'sid': self.source_id, 'runid': self.run_id,
                                        'datebegin': date_begin, 'dateend': date_end, 'unitid': unit_id,
                                        'opid': self.op_id, 'opinsid': self.op_ins_id
                                        }

                        for attribute in self.consumer.attributes:
                            # определяем известные служебные поля
                            const_value = const_values.get(consumer_roles[attribute.name])
                            if const_value:
                                const_fields_file[attribute.name] = const_value

                        if self.ctx['multiAccessMode']:
                            unique_path = self.get_unique_filepath(filepath)

                            try:
                                copyfile(filepath.path, unique_path)
                            except FileNotFoundError:
                                skip = self.file_not_found_handler(filepath, 'параллельная экстракция')
                            else:
                                filepath.path = unique_path

                        if not skip:
                            try:
                                self.reader.open_stream(filepath.path)
                            except FileNotFoundError:
                                skip = self.file_not_found_handler(filepath, 'открытие файла')
                            except Exception as e:
                                msg = f'Не удалось открыть источник {filepath.f_path}. Ошибка: {e}'
                                skip = 3
                                self.file_bogus_handler(msg)
                            else:
                                log.debug(f'Источник {filepath.f_path} успешно открыт')

                    if not skip:
                        # TODO: вынести операции с листами в ридер
                        sheets = dict()

                        if self.ctx['sheetName']:

                            if self.ctx['reInSheetNames']:
                                sheet_name_re = re.compile(self.ctx['sheetName'])
                                # ищем листы по регулярному выражению, для чего необходимо узнать какие листы есть
                                for name in self.reader.sheets:
                                    if sheet_name_re.search(name):
                                        sheets[name] = name
                                        if not self.ctx['loadMultipleSheets']:
                                            # загрузка одного листа
                                            break
                            else:
                                # лист определен по точному наименованию
                                sheets[self.ctx['sheetName']] = self.ctx['sheetName']
                        else:
                            # имя листа не задано, используем индекс
                            sheets[self.ctx['sheetOrdinal']] = str(self.ctx['sheetOrdinal'] + 1)

                        if not sheets:
                            msg = self.reader.sheet_identity_type
                            msg = f'Лист не определен. Использованный способ: {self.reader.sheet_identity_type}'
                            if self.ctx['ignoreAbsentSheets']:
                                skip = 1
                                log.warning(msg)
                            else:
                                raise DFExcelExtractorException(msg)

                        if not skip and self.reader.sheet_identity_type != 'МАСКА':
                            # точно существуют только листы, определенные по маске, остальные проверяем на наличие
                            for sheet in sheets:
                                try:
                                    self.reader.open_sheet(sheet)
                                except KeyError:
                                    msg = f'Лист не найден. {self.reader.sheet_identity}: {self.reader.sheet_str}'
                                    if self.ctx['ignoreAbsentSheets']:
                                        skip = 2
                                        log.warning(msg)
                                    else:
                                        raise DFExcelExtractorException(msg)

                    if not skip:
                        # служебные поля каждого листа, в случае отличия
                        const_fields = const_fields_file.copy()

                        for sheet in sheets:
                            # листы по маске могут быть еще не открыты
                            self.reader.open_sheet(sheet)

                            # запись информации о листе или файле в целевую сущность
                            # определяем путь и имя файла здесь, т.к. как правило, этот цикл выполняется единожды
                            for attribute in self.producer.attributes:
                                # установка окончательных для листа значений констант
                                const_value = None

                                for const_field in const_fields:
                                    if attribute.name == const_field:
                                        if const_fields_file[const_field] == 'sheetName':
                                            # если лист определяется по номеру, будет записан его номер
                                            const_value = self.get_final_const_value(attribute, self.reader.sheet_str)
                                            break
                                        elif const_fields_file[const_field] == 'fileName':
                                            if self.ctx['multiAccessMode']:
                                                base_name = filepath.filename[33:]
                                            else:
                                                base_name = filepath.filename
                                            const_value = self.get_final_const_value(attribute, base_name)
                                            break
                                        elif const_fields_file[const_field] == 'filePath':
                                            const_value = self.get_final_const_value(attribute, filepath.f_path)
                                            break

                                if const_value:
                                    const_fields[attribute.name] = const_value

                                # установка окончательных для листа значений параметров
                                # вероятно стоит добавить аннотацию позволяющаю падать на отсутствующем параметре
                                param_key = param_fields.get(attribute.name)
                                if param_key:
                                    param_fields[attribute.name] = self.reader.parameters.get(param_key)

                            # загружаем данные выбранным загрузчиком
                            total = 0
                            try:
                                total = self.load_data(rid_column, consumer_table, names_decoded, producer_columns,
                                                       const_fields, param_fields, extra_table, extras, head)
                            except FileNotFoundError:
                                skip = self.file_not_found_handler(filepath, 'загрузка данных')

                            except Exception as e:
                                msg = f'Не удалось загрузить данные из источника {filepath.f_path}. Ошибка: {e}'
                                skip = 3
                                self.file_bogus_handler(msg)
                            else:
                                log.debug(f'Данные из источника {filepath.f_path} успешно загружены')

                            self.reader.close_sheet()
                            file_total += total

                        log.info(f'Записей загружено {file_total} из {filepath.f_path} в таблицу {table_name_full}')

                        grand_total += file_total

                    if not skip:
                        processed_files[filepath] = True
                    else:
                        if skip > 0 and not self.ctx['doNotPipeSkipped']:
                            # файл пропущен, но постобработка все равно должна быть
                            processed_files[filepath] = False

                log.info(f'Всего загружено записей {grand_total} в {table_name_full}')

        except Exception as e:
            log.error(f'Ошибка экстрактора Excel: {e.__class__.__name__} {e}')
            raise e
        finally:
            self.uploader.dispose()
            self.reader.free_resources()

        # служебные операции
        if archive_folder or failed_folder:
            new_paths = set()
            for filepath in processed_files:
                path = filepath.f_path
                target_folder = archive_folder if processed_files[filepath] else failed_folder
                new_paths.add(self.archive_file(path, target_folder, base_folder,
                                                ignore_errors=self.ctx['ignoreArchiveErrors']))
            if self.ctx['removeArchivedFolders']:
                # удаляем пустые папки из переменной части пути
                processed_folders = set()
                for new_path in new_paths:
                    relative_path = os.path.relpath(new_path, start=self.escape_path(target_folder))
                    old_path = os.path.join(base_folder, relative_path)
                    parent_folder = os.path.dirname(old_path)
                    # удаляем только папки по иерархии ниже базовой папки
                    while parent_folder.startswith(base_folder + os.sep):
                        if parent_folder not in processed_folders:
                            processed_folders.add(parent_folder)
                            try:
                                log.info(f'Удаляем директорию {parent_folder}')
                                os.rmdir(parent_folder)
                            except OSError:
                                # вероятнее всего, директория не пуста, прекращаем удаление
                                log.info(f'Директория {parent_folder} не пуста, удаление отменено')
                                break
                            except Exception as e:
                                msg = f'Ошибка удаления директории {parent_folder}: {e}'
                                if self.ctx['ignoreArchiveErrors']:
                                    log.warning(msg)
                                else:
                                    raise DFExcelExtractorException(msg)
                        parent_folder = os.path.dirname(parent_folder)
